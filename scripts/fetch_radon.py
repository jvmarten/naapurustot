#!/usr/bin/env python3
"""
Fetch STUK indoor-radon concentrations for single-family houses by postal code
area and write scripts/radon.json keyed by postal code.

Data source: Säteilyturvakeskus (STUK) — "Pientalojen radonpitoisuudet
postinumeroalueittain", CC BY 4.0 (per STUK open-data policy, stuk.fi/avoin-data).
  Landing page: https://stuk.fi/pientalojen-radonpitoisuudet-postinumeroalueittain
  Serves a machine-readable .xlsx keyed by postal code. Measurements 1980–2023;
  an area is listed only if ≥10 dwellings were measured. Annual-average correction
  already applied by STUK.

We surface the per-area MEDIAN radon concentration (Bq/m³). This is genuine
postal-code-level data, so it is NOT proxy-flagged. ~1,567 postal areas have data;
the rest stay null (gray "low data" fallback per the project's partial-coverage
policy). The Finnish action level for existing dwellings is 300 Bq/m³.

Output: scripts/radon.json   {"00200": 62, "00320": 157, ...}  (median Bq/m³)

Usage:
  python scripts/fetch_radon.py
"""

import io
import json
import logging
import re
from pathlib import Path

import openpyxl
import requests

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

SCRIPT_DIR = Path(__file__).parent
OUTPUT_FILE = SCRIPT_DIR / "radon.json"

LANDING_URL = "https://stuk.fi/pientalojen-radonpitoisuudet-postinumeroalueittain"
# Fallback if the landing-page scrape fails (filename embeds the data year, so the
# scrape is preferred — it survives STUK's annual refresh to _2024 etc.).
FALLBACK_XLSX = (
    "https://stuk.fi/documents/150192312/157590338/"
    "radontilasto_pientalot_postinumero_2023.xlsx/"
    "96180b2f-6785-d7f8-8556-fe60aabff3cf?t=1712225704321"
)
HEADERS = {"User-Agent": "naapurustot-data-pipeline/1.0 (+https://naapurustot.fi)"}


def resolve_xlsx_url() -> str:
    """Scrape the postal-code .xlsx href off the STUK landing page."""
    try:
        r = requests.get(LANDING_URL, headers=HEADERS, timeout=60)
        r.raise_for_status()
        m = re.search(
            r'href="([^"]*radontilasto_pientalot_postinumero[^"]*\.xlsx[^"]*)"',
            r.text, re.IGNORECASE,
        )
        if m:
            href = m.group(1)
            if href.startswith("/"):
                href = "https://stuk.fi" + href
            logger.info("Resolved xlsx URL from landing page: %s", href.split("?")[0])
            return href
    except requests.RequestException as e:
        logger.warning("Landing-page scrape failed (%s); using fallback URL", e)
    return FALLBACK_XLSX


def fetch_radon() -> dict[str, int]:
    url = resolve_xlsx_url()
    logger.info("Downloading radon xlsx...")
    r = requests.get(url, headers=HEADERS, timeout=120)
    r.raise_for_status()

    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    header = [str(h or "").strip() for h in next(rows)]

    def col_index(predicate) -> int:
        for i, h in enumerate(header):
            if predicate(h.lower()):
                return i
        raise ValueError(f"Column not found in header: {header}")

    pno_idx = col_index(lambda h: h.startswith("postinumero") and "paikka" not in h)
    median_idx = col_index(lambda h: "mediaani" in h)

    result: dict[str, int] = {}
    for row in rows:
        if not row or row[pno_idx] is None:
            continue
        pno = str(row[pno_idx]).strip().zfill(5)
        val = row[median_idx]
        if val is None:
            continue
        try:
            result[pno] = int(round(float(val)))
        except (ValueError, TypeError):
            continue

    logger.info("Parsed median radon for %d postal codes", len(result))
    return result


def main():
    logger.info("=" * 60)
    logger.info("STUK radon (postal-code median) data pipeline")
    logger.info("=" * 60)

    data = fetch_radon()
    if not data:
        raise SystemExit("No radon data parsed — aborting without writing.")

    sorted_data = dict(sorted(data.items()))
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(sorted_data, f, indent=2, ensure_ascii=False)
    logger.info("Saved %d entries to %s", len(sorted_data), OUTPUT_FILE)

    vals = list(sorted_data.values())
    logger.info("  Range: min=%d max=%d mean=%.0f", min(vals), max(vals), sum(vals) / len(vals))


if __name__ == "__main__":
    main()
